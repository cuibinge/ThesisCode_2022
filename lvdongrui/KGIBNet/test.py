import pandas as pd
import openpyxl

xlsx_file = "/mnt/cat/CAT-Seg-main/KG/huang1.xlsx"  # 替换为实际路径
try:
    # 使用 openpyxl 检查工作表数量
    wb = openpyxl.load_workbook(xlsx_file)
    print(f"工作表数量: {len(wb.sheetnames)}")
    print(f"工作表名称: {wb.sheetnames}")
    
    # 使用 pandas 读取文件
    df = pd.read_excel(xlsx_file, header=None, usecols=[0, 1, 2])
    print("文件内容预览:")
    print(df.head())
except Exception as e:
    print(f"读取 Excel 文件时出错: {str(e)}")